"""Reading a cold-calling list out of whatever spreadsheet the owner has.

The brief is "extract numbers and names from Excel", and the hard part is that
there is no agreed shape for these files. A brokerage's calling list arrives as
whatever the last person exported: a portal dump, a purchased database, a sheet
someone typed by hand. Across real files you get

* header text that is never the same twice -- ``Name``, ``Client Name``,
  ``CUSTOMER NAME``, ``Lead Name``; ``Mobile``, ``Phone``, ``Contact No.``,
  ``Ph``, ``Cell``, ``Mobile Number``
* title rows above the header ("Andheri Leads - March", a blank row, *then* the
  headers)
* no header row at all
* phone numbers stored as Excel *numbers*, so ``9876543210`` arrives as
  ``9876543210.0`` and a leading zero is long gone
* one ``Name`` column holding "Rakesh Sharma", or separate first/last columns

So this module does not ask the owner to reformat anything. It finds the header
row, maps columns by alias, and where the aliases fail it falls back to looking
at the *values* -- a column where most cells are ten digits is a phone column
whatever its header says.

Parsing is kept separate from importing on purpose: the router previews with
these functions before writing anything, so the owner sees what was detected
and can back out rather than discovering a bad mapping after 800 rows are in
someone's queue.
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from typing import Any

from app.listing_normalize import normalize_phone

# How many rows to scan when hunting for the header.
HEADER_SEARCH_ROWS = 10

# A column qualifies as phone/name by value-sniffing if this share of its
# non-empty cells look right.
SNIFF_THRESHOLD = 0.6

MAX_ROWS = 20_000

# Header aliases, lowercased and stripped of non-letters before comparison.
# Ordered longest-first within each field so "firstname" beats "name".
FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "first_name": ("firstname", "fname", "givenname"),
    "last_name": ("lastname", "lname", "surname", "familyname"),
    "full_name": (
        "clientname",
        "customername",
        "leadname",
        "contactname",
        "partyname",
        "fullname",
        "name",
        "client",
        "customer",
    ),
    "phone": (
        "mobilenumber",
        "contactnumber",
        "phonenumber",
        "whatsappnumber",
        "mobileno",
        "contactno",
        "phoneno",
        "mobile",
        "contact",
        "phone",
        "number",
        "cell",
        "whatsapp",
        "tel",
        "ph",
    ),
    "email": ("emailaddress", "emailid", "email", "mail"),
    "budget": ("budgetrange", "budget", "price", "pricerange"),
    "location": (
        "preferredlocation",
        "preferredarea",
        "location",
        "area",
        "locality",
        "region",
        "city",
    ),
    "notes": ("remarks", "remark", "notes", "note", "comment", "comments"),
    "source": ("leadsource", "source", "channel", "campaign"),
}

_NON_ALNUM = re.compile(r"[^a-z0-9]+")
_DIGITS = re.compile(r"\d")


def _key(text: Any) -> str:
    """Normalize a header cell for alias comparison."""
    return _NON_ALNUM.sub("", str(text or "").strip().lower())


def _looks_like_phone(value: Any) -> bool:
    """Indian mobile numbers are 10 digits, often carrying +91/0 or spacing."""
    if value is None:
        return False
    digits = _DIGITS.findall(str(value))
    return 10 <= len(digits) <= 13


def _looks_like_name(value: Any) -> bool:
    text = str(value or "").strip()
    if not text or len(text) > 60:
        return False
    letters = sum(c.isalpha() or c.isspace() for c in text)
    # Mostly letters, and not a number that happens to be stored as text.
    return letters / max(len(text), 1) > 0.8 and any(c.isalpha() for c in text)


def _cell(value: Any) -> str:
    """Render a spreadsheet cell as clean text.

    Excel stores a phone number typed without formatting as a float, so
    ``9876543210`` round-trips as ``9876543210.0``. Trim that here rather than
    letting it reach the phone normalizer, which would read the trailing
    ``.0`` as two more digits.
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


@dataclass
class ParsedRow:
    row_number: int
    first_name: str = ""
    last_name: str = ""
    phone: str = ""
    email: str = ""
    location: str = ""
    notes: str = ""
    source: str = ""
    # Why this row cannot be imported, if it cannot be.
    problem: str = ""

    @property
    def usable(self) -> bool:
        return not self.problem

    @property
    def display_name(self) -> str:
        return f"{self.first_name} {self.last_name}".strip()


@dataclass
class ParseResult:
    rows: list[ParsedRow] = field(default_factory=list)
    # Which spreadsheet column ended up mapped to which field, for the preview.
    detected_columns: dict[str, str] = field(default_factory=dict)
    header_row: int | None = None
    total_rows: int = 0
    sheet_name: str | None = None
    warnings: list[str] = field(default_factory=list)

    @property
    def usable_rows(self) -> list[ParsedRow]:
        return [r for r in self.rows if r.usable]


def _read_table(data: bytes, filename: str) -> tuple[list[list[Any]], str | None]:
    """Return the sheet as a list of rows, plus the sheet name for xlsx."""
    lowered = (filename or "").lower()

    if lowered.endswith((".xlsx", ".xlsm")):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - dependency is pinned
            raise ValueError(
                "Excel support needs openpyxl. Run: pip install -r requirements.txt"
            ) from exc
        # read_only keeps a large sheet from being materialised twice; data_only
        # returns the cached result of a formula rather than "=A1&B1".
        workbook = load_workbook(
            io.BytesIO(data), read_only=True, data_only=True
        )
        sheet = workbook.active
        rows = [list(r) for r in sheet.iter_rows(values_only=True, max_row=MAX_ROWS)]
        name = sheet.title
        workbook.close()
        return rows, name

    if lowered.endswith(".xls"):
        raise ValueError(
            "Legacy .xls files are not supported. Open it in Excel or Sheets "
            "and re-save as .xlsx or .csv."
        )

    # CSV / TSV. utf-8-sig strips the BOM Excel writes on 'CSV UTF-8' export.
    text = data.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    return [list(r) for r in csv.reader(io.StringIO(text), dialect)], None


def _find_header(rows: list[list[Any]]) -> tuple[int | None, dict[str, int]]:
    """Locate the header row and map field -> column index.

    Scores each candidate row by how many known aliases it matches, so a title
    row ("Andheri Leads — March") loses to the real header two rows below it.
    """
    best: tuple[int, dict[str, int]] | None = None

    for index, row in enumerate(rows[:HEADER_SEARCH_ROWS]):
        mapping: dict[str, int] = {}
        for col, cell in enumerate(row):
            key = _key(cell)
            if not key:
                continue
            for field_name, aliases in FIELD_ALIASES.items():
                if field_name in mapping:
                    continue
                if any(key == alias for alias in aliases):
                    mapping[field_name] = col
                    break
        if mapping and (best is None or len(mapping) > len(best[1])):
            best = (index, mapping)

    if best is None:
        return None, {}
    return best[0], best[1]


def _sniff_columns(
    rows: list[list[Any]], start: int, mapping: dict[str, int]
) -> dict[str, int]:
    """Fill gaps by inspecting values when the header didn't name them.

    This is what rescues a file with no header row, or one whose phone column
    is titled something nobody anticipated.
    """
    sample = rows[start : start + 60]
    if not sample:
        return mapping

    width = max((len(r) for r in sample), default=0)
    taken = set(mapping.values())

    for col in range(width):
        values = [r[col] for r in sample if col < len(r) and _cell(r[col])]
        if not values:
            continue

        if "phone" not in mapping and col not in taken:
            hits = sum(_looks_like_phone(v) for v in values)
            if hits / len(values) >= SNIFF_THRESHOLD:
                mapping["phone"] = col
                taken.add(col)
                continue

        if (
            "full_name" not in mapping
            and "first_name" not in mapping
            and col not in taken
        ):
            hits = sum(_looks_like_name(v) for v in values)
            if hits / len(values) >= SNIFF_THRESHOLD:
                mapping["full_name"] = col
                taken.add(col)

    return mapping


def _split_name(full: str) -> tuple[str, str]:
    """Split a single name cell into first and last.

    Everything after the first token becomes the surname, so "Rakesh Kumar
    Sharma" keeps "Kumar Sharma" together rather than dropping the middle name.
    """
    parts = [p for p in re.split(r"\s+", full.strip()) if p]
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def parse_lead_file(data: bytes, filename: str) -> ParseResult:
    """Parse a spreadsheet into rows ready for import review."""
    rows, sheet_name = _read_table(data, filename)
    result = ParseResult(sheet_name=sheet_name)

    rows = [r for r in rows if any(_cell(c) for c in r)]  # drop blank rows
    if not rows:
        result.warnings.append("The file has no data in it.")
        return result

    header_index, mapping = _find_header(rows)
    data_start = (header_index + 1) if header_index is not None else 0
    result.header_row = (header_index + 1) if header_index is not None else None

    mapping = _sniff_columns(rows, data_start, mapping)

    if "phone" not in mapping:
        result.warnings.append(
            "No phone column found. A calling list without numbers cannot be "
            "worked, so nothing will be imported."
        )
    if not {"full_name", "first_name"} & set(mapping):
        result.warnings.append(
            "No name column found — rows will import as 'Unknown' and are "
            "worth checking before you assign them."
        )

    header_cells = rows[header_index] if header_index is not None else []
    result.detected_columns = {
        field_name: (
            _cell(header_cells[col])
            if col < len(header_cells) and _cell(header_cells[col])
            else f"column {col + 1}"
        )
        for field_name, col in sorted(mapping.items(), key=lambda kv: kv[1])
    }

    def value(row: list[Any], key: str) -> str:
        col = mapping.get(key)
        if col is None or col >= len(row):
            return ""
        return _cell(row[col])

    seen_phones: set[str] = set()

    for offset, row in enumerate(rows[data_start:]):
        line = data_start + offset + 1
        parsed = ParsedRow(row_number=line)

        if "first_name" in mapping:
            parsed.first_name = value(row, "first_name")
            parsed.last_name = value(row, "last_name")
        else:
            parsed.first_name, parsed.last_name = _split_name(value(row, "full_name"))

        raw_phone = value(row, "phone")
        parsed.phone = normalize_phone(raw_phone) or ""
        parsed.email = value(row, "email")
        parsed.location = value(row, "location")
        parsed.notes = value(row, "notes")
        parsed.source = value(row, "source")

        if not parsed.phone:
            parsed.problem = (
                f"no usable phone number ({raw_phone!r})" if raw_phone else "no phone number"
            )
        elif parsed.phone in seen_phones:
            # Same number twice in one file — a merged export, usually.
            parsed.problem = "duplicate of an earlier row in this file"
        else:
            seen_phones.add(parsed.phone)
            if not parsed.first_name:
                parsed.first_name = "Unknown"

        result.rows.append(parsed)

    result.total_rows = len(result.rows)
    return result


def distribute(items: list[Any], buckets: int) -> list[list[Any]]:
    """Deal items round-robin, like a deck of cards.

    Round-robin rather than contiguous slices so that a file sorted by area or
    by budget does not hand one caller every premium lead and another the
    leftovers.
    """
    if buckets <= 0:
        return []
    out: list[list[Any]] = [[] for _ in range(buckets)]
    for index, item in enumerate(items):
        out[index % buckets].append(item)
    return out
